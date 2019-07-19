from openpyxl import load_workbook  # load_workbook 是一个函数  来进行excel文件的读写

class Case:
    #定义存储测试数据的类
    def __init__(self):
        self.id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None

class DoExcel:
    #定义读excel的一个类
    def __init__(self,filename):
        try:
            # 操作的文件
            self.filename = filename
            # 实例化一个workbook对象
            self.workbook = load_workbook(filename = filename)
        except FileNotFoundError as e:
            # 文件未找到异常处理
            print('{0} not found,please check file path'.format(self.filename))
            raise e

    def get_data(self,sheet_name):
        #读取数据
        sheet_name = sheet_name
        sheet = self.workbook[sheet_name]  #根据表单名称获取表单（定位表单）

        case = []#所有的数据存放在大列表中
        for i in range(2,sheet.max_row+1):
            #因为下标从1开始，所以我需要从2开始读
            #sheet.max_row+1 获取最大行
            cases = Case() #每一行数据存在对象Cases()里面，每一行数据都相当于case的实例
            #分别按照每行每列的值取出来，赋值到对应case的属性里面
            cases.id = sheet.cell(row = i, column = 1).value #取第r行，第1格的值
            cases.title = sheet.cell(row = i,column = 2).value #取第r行，第2格的值
            cases.url = sheet.cell(row = i, column = 3).value #取第r行，第3格的值
            cases.data = sheet.cell(row = i, column = 4).value #取第r行，第4格的值
            cases.method = sheet.cell(row = i, column = 5).value #取第r行，第5格的值
            cases.expected = sheet.cell(row = i, column = 6).value #取第r行，第6格的值
            case.append(cases)
        return case

    def write_result(self,sheet_name,row,actual,result):
        #写回数据
        sheet = self.workbook[sheet_name]#定位表单
        sheet.cell(row,7).value = actual # 写入实际结果
        sheet.cell(row,8).value = result # 写入执行结果，PASS or FAIL
        #写入后保存
        self.workbook.save(filename = self.filename)


if __name__ == '__main__':
    from common import contants
    from common.request import Request

    # import json #json 是一个库，所以需要import 进来
    # people = '{"name":"lili","age":18,"married":false,"remark":null}' #json格式的字符串
    # print(people)
    # o_dict = json.loads(people) #json格式字符串序列化为字典
    # print(type(o_dict),o_dict)
    # print(type(eval(o_dict)),o_dict) #eval函数来尝试把上面的那个字符串变成字典，不行。eval是函数，是python的函数，是基于python的格式进行转换的
    # 如何有100个参数如何快速的构造入参（特别是后台的接口，参数比较多）
    # 把入参保存入一份文件中，把它序列化出来变成一个字典，如放入data.json的一个file
    # 用json.load()

    # fp = open(contants.data_json) #读取文件
    # f_dict = json.load(fp=fp)
    # print(f_dict['name'])

    do_excel = DoExcel(contants.case_file)
    rest = do_excel.get_data('login')

    request = Request() #实例化对象
    for item in rest :
        #参数处理
        a=item.__dict__ #用来辅助debug的  可以查看对象里面的内容
        print(type(item.data),item.data) #从excel里面读出来的数据是一个个的字符串，所以需要转换
        resp = request.request(item.method,item.url,item.data)
        print(resp.json())
        # print(type(resp.json()),resp.json())
        if resp.text == item.expected:
            do_excel.write_result('login',item.id+1,resp.text,'PASS') #item.id+1是当前的row，resp.text真实值
        else:
            do_excel.write_result('login',item.id+1,resp.text,'FAIL')






